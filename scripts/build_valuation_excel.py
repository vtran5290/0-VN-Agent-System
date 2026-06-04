"""
Build Vietnam Market Valuation Verification Excel Workbook
Date: 2026-05-30
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')

OUTPUT_PATH = "data/decision/vietnam_market_valuation_verification_20260530.xlsx"

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUB_FILL = PatternFill("solid", fgColor="2E75B6")
CONFIRM_FILL = PatternFill("solid", fgColor="C6EFCE")
REVISED_FILL = PatternFill("solid", fgColor="FFEB9C")
REJECT_FILL = PatternFill("solid", fgColor="FFC7CE")
UNVERIF_FILL = PatternFill("solid", fgColor="EDEDED")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
BOLD = Font(bold=True)
ITALIC = Font(italic=True)

def style_header(cell, fill=HEADER_FILL):
    cell.font = HDR_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_sub(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = SUB_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')

def auto_width(ws, min_width=10, max_width=45):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)

def status_fill(status):
    s = str(status).upper()
    if 'CONFIRMED' in s:
        return CONFIRM_FILL
    elif 'REVISED' in s:
        return REVISED_FILL
    elif 'REJECTED' in s or 'REJECT' in s:
        return REJECT_FILL
    else:
        return UNVERIF_FILL

# -----------------------------------------------------------------------
# TAB 0: README
# -----------------------------------------------------------------------
ws0 = wb.active
ws0.title = "00_ReadMe_Methodology"
readme = [
    ("Vietnam Market & Sector Valuation Verification", None),
    (f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", None),
    ("", None),
    ("DATA SOURCES", None),
    ("Primary", "FireAnt SSOT — fa_quarterly.parquet (53,395 rows, 1,932 symbols, through 2026Q2)"),
    ("Prices", "FireAnt SSOT — ta_ohlcv_panel.parquet (max date 2026-05-29)"),
    ("VN-Index", "FireAnt SSOT — ta_vnindex.parquet (2012-01-03 to 2026-05-15)"),
    ("Market Cap / Sector", "TradingView screener (top 200 Vietnamese stocks, 2026-05-29)"),
    ("Macro", "FRED (UST rates, CPI), SBV (OMO, interbank, FX), local weekly report"),
    ("", None),
    ("METHODOLOGY", None),
    ("P/E", "Market-cap weighted: Sum(MC) / Sum(LTM profit) for positive earners only"),
    ("P/B", "Market-cap weighted: Sum(MC) / Sum(book equity via BVPS x shares)"),
    ("LTM Profit", "financialValues_ParentCompanyShareholderProfitAfterTax_TTM from FA quarterly"),
    ("Book Equity", "BVPS (VND/share) x shares outstanding; fallback to TotalShareHolderEquity"),
    ("Market Cap", "TradingView tv_mc (USD); verified vs FA MarketCapAtPeriodEnd"),
    ("Exchange Rate", "25,137 VND/USD (SBV reference 2026-05-29)"),
    ("Historical bottoms", "Price minimum in 5-7 day window around bottom date x concurrent FA quarter"),
    ("", None),
    ("COVERAGE NOTES", None),
    ("Index proxy", "Top 200 stocks by market cap (TradingView) covers ~99.8% of reported mktcap"),
    ("Negative earners", "Excluded from P/E numerator; included in mktcap denominator"),
    ("VIC", "P/E null — earnings near breakeven; P/B 10.76x reflects goodwill/conglomerate premium"),
    ("VPL", "Very recent listing; no FA TTM history; P/E and P/B not available"),
    ("HVN", "P/E 7.5x but P/B 11.2x and ROE 446% — distorted by post-restructuring balance sheet"),
    ("Forward P/E", "INFERRED from LTM + SSI/broker consensus EPS growth; not directly computed"),
    ("2009/2012/2016 bottoms", "FA data starts 2016; cannot compute from local data"),
]
ws0.column_dimensions['A'].width = 30
ws0.column_dimensions['B'].width = 65
for i, (k, v) in enumerate(readme, 1):
    ws0.cell(i, 1, k)
    if v:
        ws0.cell(i, 2, v)
    if k and not v and k != "":
        ws0.cell(i, 1).font = BOLD
        ws0.cell(i, 1).fill = SUB_FILL
        ws0.cell(i, 1).font = Font(bold=True, color="FFFFFF")
        ws0.merge_cells(f'A{i}:B{i}')

# -----------------------------------------------------------------------
# TAB 1: INDEX VALUATION
# -----------------------------------------------------------------------
ws1 = wb.create_sheet("01_Index_Valuation")
headers = ["Version", "N Stocks", "Mktcap $B", "LTM P/E", "LTM P/B", "Neg Earners", "Notes"]
for c, h in enumerate(headers, 1):
    cell = ws1.cell(1, c, h)
    style_header(cell)

rows = [
    ("Full Market (Top 200, cap-wtd)", 200, 352.6, 13.25, 2.00, 2, "Primary estimate; 99.8% mktcap coverage"),
    ("Full Market (All 1477 stocks)", 1477, "~404", 14.40, 2.09, 103, "Includes micro-cap distortions; not VN-Index proxy"),
    ("Ex VIC + VHM", 198, 267.9, 11.39, 1.68, 2, "Remove only VHM (10.0x PE) and VIC (null PE)"),
    ("Ex Vin Group (VIC+VHM+VRE+VPL)", 196, 258.9, 11.19, 1.66, 2, "Full Vin group exclusion"),
    ("Vin Group Only", 4, 93.7, 27.04, 4.58, 0, "VIC near-breakeven earnings inflating agg P/E"),
]
for r, row in enumerate(rows, 2):
    for c, val in enumerate(row, 1):
        ws1.cell(r, c, val)
    ws1.cell(r, 1).font = BOLD

ws1.append([])
ws1.append(["CHATGPT CLAIM vs VERIFICATION", "", "", "", "", "", ""])
ws1.cell(ws1.max_row, 1).font = BOLD
ws1.cell(ws1.max_row, 1).fill = SUB_FILL
ws1.cell(ws1.max_row, 1).font = Font(bold=True, color="FFFFFF")

claims_hdr = ["Claim", "ChatGPT Value", "Our Calculation", "Delta", "Status", "Comment"]
ws1.append(claims_hdr)
for c in range(1, 7):
    style_header(ws1.cell(ws1.max_row, c), fill=SUB_FILL)

claims = [
    ("VN-Index LTM P/E", "13.6-13.8x", "13.25x", "-0.35-0.55x", "REVISED", "Slightly lower; timing/coverage difference"),
    ("VN-Index LTM P/B", "2.08x", "2.00-2.09x", "-0.01 to +0.01", "CONFIRMED", "Range straddles our two estimates"),
    ("Ex VIC+VHM LTM P/E", "11.6x", "11.39x", "-0.21x", "REVISED", "Lower than claimed"),
    ("Ex Vin Group LTM P/E", "11.6x", "11.19x", "-0.41x", "REVISED", "Lower than claimed"),
    ("Ex Vin Group LTM P/B", "1.75x", "1.66x", "-0.09x", "REVISED", "Lower by 5%"),
    ("Fwd 2026 P/E headline", "11.5-11.7x", "~11.6x (implied)", "~0", "APPROX CONSISTENT", "Cannot directly verify; consistent with SSI target"),
    ("Fwd 2026 P/E ex-Vin", "9.8-10.1x", "~9.9x (implied)", "~0", "APPROX CONSISTENT", "Implied from LTM + 13% earnings growth ex-Vin"),
]
for row in claims:
    ws1.append(list(row))
    r = ws1.max_row
    fill = status_fill(row[4])
    for c in range(1, 7):
        ws1.cell(r, c).fill = fill
        ws1.cell(r, c).alignment = Alignment(wrap_text=True)

auto_width(ws1)
ws1.row_dimensions[1].height = 30

# -----------------------------------------------------------------------
# TAB 2: EX-VIN CALCULATION
# -----------------------------------------------------------------------
ws2 = wb.create_sheet("02_Ex_Vin_Calculation")
ws2.append(["Vin Group Breakdown", "", "", "", "", ""])
ws2.cell(1,1).font = BOLD
ws2.cell(1,1).fill = HEADER_FILL
ws2.cell(1,1).font = HDR_FONT

hdrs = ["Ticker", "Mktcap $B", "LTM Profit $M", "Book Equity $B", "P/E (TTM)", "P/B", "ROE%", "Exchange"]
ws2.append(hdrs)
for c in range(1, 9):
    style_sub(ws2.cell(2, c))

vin_data = [
    ("VIC", 60.1, 547, 6.1, "null (breakeven)", 10.76, 9.6, "HOSE"),
    ("VHM", 24.6, 2545, 10.9, 10.0, 2.69, 27.4, "HOSE"),
    ("VRE", 2.8, 273, 2.0, 10.7, 1.52, 14.8, "HOSE"),
    ("VPL", 6.3, 101, 1.5, "null (new listing)", "null", "null", "HOSE"),
    ("TOTAL", 93.7, 3466, 20.5, 27.04, 4.58, "14.8 avg", "—"),
]
for row in vin_data:
    ws2.append(list(row))

ws2.append([])
ws2.append(["Impact of Vin removal on index P/E", "", "", "", "", "", "", ""])
ws2.cell(ws2.max_row, 1).font = BOLD

impact_hdrs = ["Scenario", "Mktcap $B", "Mktcap % of Full", "Agg LTM P/E", "Agg LTM P/B", "Delta P/E vs Full", "Delta P/B vs Full"]
ws2.append(impact_hdrs)
for c in range(1, 8):
    style_sub(ws2.cell(ws2.max_row, c))

impact_data = [
    ("Full Market", 352.6, "100%", 13.25, 2.00, "—", "—"),
    ("Ex VIC + VHM", 267.9, "76.0%", 11.39, 1.68, "-1.86x", "-0.32x"),
    ("Ex Vin Group", 258.9, "73.4%", 11.19, 1.66, "-2.06x", "-0.34x"),
    ("Vin Group Only", 93.7, "26.6%", 27.04, 4.58, "+13.79x vs ex-Vin", "+2.92x vs ex-Vin"),
]
for row in impact_data:
    ws2.append(list(row))

auto_width(ws2)

# -----------------------------------------------------------------------
# TAB 3: CYCLE BOTTOMS
# -----------------------------------------------------------------------
ws3 = wb.create_sheet("03_Cycle_Bottoms")
ws3.append(["VN-Index Historical Cycle Bottoms — P/E and P/B Verification", "", "", "", "", "", "", ""])
ws3.cell(1,1).font = HDR_FONT
ws3.cell(1,1).fill = HEADER_FILL
ws3.merge_cells('A1:H1')

hdrs3 = ["Period", "VNI Level", "Our P/E", "ChatGPT P/E", "P/E Status", "Our P/B", "ChatGPT P/B", "P/B Status", "Data Basis"]
ws3.append(hdrs3)
for c in range(1, 10):
    style_sub(ws3.cell(2, c))

bottoms = [
    ("2009 bottom (Jun)", 235, "UNVERIFIABLE", 10.46, "UNVERIFIABLE", "UNVERIFIABLE", 1.24, "UNVERIFIABLE", "FA data starts 2016; cannot compute"),
    ("2012 bottom (Nov)", 337, "UNVERIFIABLE", 7.41, "UNVERIFIABLE", "UNVERIFIABLE", 1.25, "UNVERIFIABLE", "FA data starts 2016; cannot compute"),
    ("2016 bottom (Jan)", 522, "UNVERIFIABLE", 12.71, "UNVERIFIABLE", "UNVERIFIABLE", 1.78, "UNVERIFIABLE", "FA 2015Q3 not available in SSOT"),
    ("2020 Covid (Mar-19)", 659, 9.58, 10.45, "REVISED ↓", 1.49, 1.65, "REVISED ↓", "FA 2019Q4 x OHLCV Mar 2020; N=230 stocks"),
    ("2022 crisis (Nov-11)", 912, 8.74, 9.98, "REVISED ↓", 1.45, 1.71, "REVISED ↓", "FA 2022Q3 x OHLCV Nov 2022; N=359 stocks"),
    ("Current (May-29, 2026)", 1863, 13.25, "N/A", "NEW DATA", 2.00, "N/A", "NEW DATA", "FA 2025Q4 x TV mktcap; N=200 stocks (top 200)"),
]
for row in bottoms:
    ws3.append(list(row))
    r = ws3.max_row
    pe_status = str(row[4])
    pb_status = str(row[7])
    for c in [5]:
        ws3.cell(r, c).fill = status_fill(pe_status)
    for c in [8]:
        ws3.cell(r, c).fill = status_fill(pb_status)

ws3.append([])
ws3.append(["KEY FINDING:", "ChatGPT overstated cycle bottom P/E by ~10-15% and P/B by ~15-18%. Current market is NOT near 2020/2022 bottom-zone by our verified figures."])
ws3.cell(ws3.max_row, 1).font = BOLD
ws3.cell(ws3.max_row, 2).fill = REJECT_FILL
ws3.cell(ws3.max_row, 2).font = BOLD

auto_width(ws3)

# -----------------------------------------------------------------------
# TAB 4: SECTOR VALUATION
# -----------------------------------------------------------------------
ws4 = wb.create_sheet("04_Sector_Valuation")
ws4.append(["Vietnam Sector Valuation — May 2026", "", "", "", "", "", "", "", "", ""])
ws4.cell(1,1).font = HDR_FONT
ws4.cell(1,1).fill = HEADER_FILL
ws4.merge_cells('A1:J1')

hdrs4 = ["Sector", "N Stocks", "Mktcap $B", "Agg LTM P/E", "Agg LTM P/B", "Med P/E", "Med P/B", "Med ROE%", "Neg Earners", "Top 5 Tickers"]
ws4.append(hdrs4)
for c in range(1, 11):
    style_sub(ws4.cell(2, c))

sector_data = [
    ("Banks", 23, 103.0, 9.08, 1.42, 8.02, 1.19, 17.0, 0, "VCB, BID, CTG, TCB, VPB"),
    ("Securities & Fin Svcs", 15, 10.6, 13.65, 1.59, 14.72, 1.50, 10.1, 0, "SSI, VIX, HCM, VCI, VND"),
    ("Real Estate (Vin group)", 4, 93.7, 27.04, 4.58, 10.33, 2.69, 14.8, 0, "VIC, VHM, VPL, VRE"),
    ("Real Estate (ex-Vin)", 12, 8.2, 17.61, 1.07, 21.02, 1.27, 5.9, 0, "BCM, NVL, KDH, DXG, PDR"),
    ("Industrial Parks", 6, 4.9, 14.02, 1.63, 13.51, 2.35, 16.3, 0, "BCM, KBC, VGC, IDC, SZC"),
    ("Insurance", 6, 3.2, 13.27, 1.69, 11.19, 1.63, 13.2, 0, "BVH, PVI, BIC, VNR, MIG"),
    ("Consumer Goods & F&B", 11, 20.3, 15.1, 2.99, 13.4, 2.02, 16.5, 0, "MCH, VNM, MSN, SAB, KDC"),
    ("Oil & Gas", 6, 16.6, 13.6, 1.95, 12.9, 1.74, 14.0, 0, "GAS, BSR, PLX, PVD, PVS"),
    ("Chemicals & Fertilizer", 5, 7.5, 14.4, 1.78, 10.3, 2.29, 14.5, 1, "GVR, DCM, DGC, DPM, SBT"),
    ("Steel & Materials", 4, 8.7, 9.5, 1.36, 13.9, 1.10, 9.5, 0, "HPG, HSG, NKG, PHR"),
    ("Power & Utilities", 9, 11.8, 13.8, 1.74, 10.6, 1.55, 17.6, 0, "GAS, POW, REE, GEG, PC1"),
    ("Transport & Logistics", 7, 3.0, 11.9, 1.47, 14.6, 1.48, 16.5, 0, "GMD, VTP, PVT, HAH, PHP"),
    ("Aviation", 4, 12.5, 13.2, 2.82, 23.8, 2.62, 11.4, 0, "ACV, VJC, HVN, SAS"),
    ("Technology & Telecom", 5, 10.1, 14.4, 3.81, 15.9, 3.34, 20.1, 0, "FPT, GEE, CMG, FOX, ELC"),
    ("Industrial Goods/Elec", 3, 3.2, 15.2, 1.17, 17.1, 2.36, 9.2, 0, "GEX, REE, PC1"),
    ("Retail", 1, 4.3, 13.3, 3.03, 13.7, 3.44, 25.4, 0, "MWG"),
    ("Seafood Export", 2, 0.7, 6.9, 1.22, 7.5, 1.22, 15.0, 0, "VHC, ANV"),
    ("Textile Export", 1, 0.1, 5.8, 1.61, 6.2, 1.61, 12.0, 0, "MSH, TNG, TCM"),
]
for row in sector_data:
    ws4.append(list(row))

auto_width(ws4)

# -----------------------------------------------------------------------
# TAB 5: SECTOR RANKING
# -----------------------------------------------------------------------
ws5 = wb.create_sheet("05_Sector_Ranking")
ws5.append(["Sector Investment Ranking — May 2026", "", "", "", "", "", "", ""])
ws5.cell(1,1).font = HDR_FONT
ws5.cell(1,1).fill = HEADER_FILL
ws5.merge_cells('A1:H1')

hdrs5 = ["Rank", "Sector", "Valuation", "Earnings Visibility", "Balance Sheet", "Liquidity", "Catalyst", "Key Risk"]
ws5.append(hdrs5)
for c in range(1, 9):
    style_sub(ws5.cell(2, c))

ranking = [
    (1, "Banks (CTG/MBB/ACB/BID)", "CHEAP (P/E 6-9x)", "HIGH (credit growth structural)", "SOLID", "HIGH", "FTSE upgrade + credit cycle", "Credit quality / interest rate"),
    (2, "HPG (Steel)", "CHEAP (P/E 9.2x)", "MEDIUM-HIGH (Dung Quat 2 ramp)", "GOOD", "HIGH", "Dung Quat 2 capacity 2026-27", "China dumping / construction slowdown"),
    (3, "FPT (Tech)", "FAIR (P/E 12.6x)", "HIGH (tech outsourcing)", "STRONG", "MEDIUM-HIGH", "AI/outsourcing demand secular", "Earnings growth deceleration"),
    (4, "DGC (Chemicals)", "CHEAP (P/E 7.1x)", "MEDIUM (phosphate chemicals)", "GOOD", "MEDIUM", "Specialty chemicals demand", "Input cost / cyclical peak"),
    (5, "DCM (Fertilizer)", "FAIR (P/E 10.4x)", "MEDIUM (gas supply)", "GOOD", "MEDIUM", "Agricultural demand", "Gas price spike"),
    (6, "VHC (Seafood)", "CHEAP (P/E 7-8x LTM)", "MEDIUM (tariff uncertainty)", "GOOD", "LOW-MEDIUM", "US tariff de-escalation", "US tariff / demand softness"),
    (7, "VCB (Bank — premium)", "FAIR (P/E 14.4x)", "HIGH (quality)", "STRONGEST", "HIGH", "FTSE upgrade passive inflow", "Premium valuation vs peers"),
    (8, "MWG (Retail)", "FAIR (P/E 13.7x)", "MEDIUM-HIGH (recovery)", "OK", "HIGH", "Consumer spending recovery", "Competition from groceries"),
    (9, "PVS (Oilfield)", "FAIR (P/E 10.5x)", "MEDIUM (oil capex)", "GOOD", "MEDIUM", "Block B gas project", "Oil price / project delay"),
    (10, "TNG/MSH (Textile)", "CHEAP (P/E 6-8x)", "LOW-MEDIUM", "OK", "LOW", "Post-tariff order recovery", "US tariff cliff; order visibility"),
    ("AVOID", "NVL (Real Estate)", "TRAP (P/E 12x, P/B 0.74x)", "LOW (legal delays)", "WEAK", "MEDIUM", "None clear", "Project legality; balance sheet"),
    ("AVOID", "PLX (Oil Dist)", "EXPENSIVE (P/E 20.2x)", "LOW", "OK", "MEDIUM-HIGH", "None", "Margin compression"),
    ("AVOID", "HVN (Vietnam Air)", "DISTORTED (P/B 11.2x)", "LOW", "VERY WEAK", "LOW", "Oil price decline", "Balance sheet restructuring ongoing"),
    ("AVOID", "Industrial Parks (BCM,KBC)", "FAIR-EXPENSIVE (P/E 14-17x)", "LUMPY", "OK", "LOW-MEDIUM", "FDI land handover", "Lumpy revenue; execution risk"),
]
for row in ranking:
    ws5.append(list(row))
    r = ws5.max_row
    if str(row[0]) == "AVOID":
        for c in range(1, 9):
            ws5.cell(r, c).fill = REJECT_FILL
    elif row[0] <= 3:
        for c in range(1, 9):
            ws5.cell(r, c).fill = CONFIRM_FILL
    elif row[0] <= 6:
        for c in range(1, 9):
            ws5.cell(r, c).fill = REVISED_FILL

auto_width(ws5)

# -----------------------------------------------------------------------
# TAB 6: LIQUIDITY FILTER
# -----------------------------------------------------------------------
ws6 = wb.create_sheet("06_Liquidity_Filter")
ws6.append(["Liquidity Screen — ADTV (TradingView 10-day average)", "", "", "", "", ""])
ws6.cell(1,1).font = HDR_FONT
ws6.cell(1,1).fill = HEADER_FILL
ws6.merge_cells('A1:F1')

hdrs6 = ["Ticker", "Sector", "Mktcap $B", "Price (VND)", "ADTV 10D (shares)", "Liquidity Tier"]
ws6.append(hdrs6)
for c in range(1, 7):
    style_sub(ws6.cell(2, c))

# Key stocks from TradingView data (ADTV in shares; approx VND bn = adtv_shares * price * 1000 / 1e9)
liquidity_data = [
    ("MSB", "Finance/Banks", 1.79, 15300, 55607979, "HIGH (>100bn/day)"),
    ("SHB", "Finance/Banks", 2.65, 13800, 61489768, "HIGH"),
    ("VIX", "Finance/Securities", 1.64, 17700, 48011024, "HIGH"),
    ("VPB", "Finance/Banks", 8.23, 27100, 18278400, "HIGH"),
    ("MBB", "Finance/Banks", 7.65, 25000, 19553196, "HIGH"),
    ("ACB", "Finance/Banks", 4.81, 24900, 25801369, "HIGH"),
    ("HPG", "Non-Energy Minerals", 7.70, 24000, 27554190, "HIGH"),
    ("TCB", "Finance/Banks", 8.78, 32750, 14205048, "HIGH"),
    ("VCB", "Finance/Banks", 19.94, 62000, 11439562, "HIGH"),
    ("CTG", "Finance/Banks", 10.29, 34800, 10526805, "HIGH"),
    ("FPT", "Technology Services", 4.59, 71600, 12923320, "HIGH"),
    ("SSI", "Finance/Securities", 2.61, 27500, 18700079, "HIGH"),
    ("HDB", "Finance/Banks", 4.92, 25900, 21691406, "HIGH"),
    ("VND", "Finance/Securities", 1.03, 16850, 22185305, "HIGH"),
    ("GAS", "Utilities", 7.49, 87400, 2740823, "MEDIUM"),
    ("VNM", "Consumer Non-Durables", 4.69, 59200, 3687782, "MEDIUM"),
    ("MWG", "Retail Trade", 4.34, 76300, 7737646, "MEDIUM-HIGH"),
    ("BID", "Finance/Banks", 11.76, 42000, 8825693, "HIGH"),
    ("BSR", "Energy Minerals", 5.42, 29750, 15927260, "HIGH"),
    ("VJC", "Transportation", 3.83, 171900, 5308560, "MEDIUM"),
    ("GMD", "Transportation", 1.17, 72300, 929692, "LOW-MEDIUM"),
    ("POW", "Utilities", 1.61, 13700, 11744625, "HIGH"),
    ("GEX", "Electronic Technology", 1.61, 32100, 16839063, "HIGH"),
    ("PVD", "Energy Minerals", 0.63, 30500, 6196061, "MEDIUM"),
    ("PVS", "Industrial Services", 0.75, 39000, 4624558, "MEDIUM"),
    ("DCM", "Process Industries", 0.81, 39800, 2797888, "MEDIUM"),
    ("DPM", "Process Industries", 0.67, 25900, 3572225, "MEDIUM"),
    ("DGC", "Process Industries", 0.69, 46500, 1934873, "MEDIUM"),
    ("VHC", "Process Industries", 0.37, 65000, 500000, "LOW (est.)"),
    ("NVL", "Finance/RE", 1.28, 15100, 19693393, "HIGH (but value trap)"),
]
for row in liquidity_data:
    ws6.append(list(row))
    r = ws6.max_row
    tier = str(row[5]).upper()
    if "HIGH" in tier and "LOW" not in tier and "MEDIUM" not in tier and "TRAP" not in tier:
        ws6.cell(r, 6).fill = CONFIRM_FILL
    elif "MEDIUM" in tier:
        ws6.cell(r, 6).fill = REVISED_FILL
    else:
        ws6.cell(r, 6).fill = REJECT_FILL

auto_width(ws6)

# -----------------------------------------------------------------------
# TAB 7: MACRO REGIME
# -----------------------------------------------------------------------
ws7 = wb.create_sheet("07_Macro_Regime")
ws7.append(["Macro Regime Assessment — May 2026-05-29", "", "", "", ""])
ws7.cell(1,1).font = HDR_FONT
ws7.cell(1,1).fill = HEADER_FILL
ws7.merge_cells('A1:E1')

macro_hdr = ["Indicator", "Value", "Signal", "Direction", "Source"]
ws7.append(macro_hdr)
for c in range(1, 6):
    style_sub(ws7.cell(2, c))

macro_data = [
    ("VN-Index", "1,863.67", "Neutral (mid-range)", "Declining from ~1,970 high", "Weekly report 2026-05-29"),
    ("VN30", "1,999.82", "Below MA20 = NEGATIVE", "Weakening", "Weekly report"),
    ("Dist. Days (rolling 20)", "5", "ELEVATED (>4 = caution)", "Increasing trend", "VN30 proxy"),
    ("HNX above MA20", "TRUE", "Still positive", "Holding", "Weekly report"),
    ("UPCOM above MA20", "FALSE", "Breadth weakening", "Declining", "Weekly report"),
    ("UST 10Y", "4.47%", "TIGHT (+20bp WoW)", "Rising = risk-off", "FRED DGS10"),
    ("UST 2Y", "4.00%", "Mild inversion resolved", "+24bp WoW — rate volatility", "FRED DGS2"),
    ("USD broad (DTWEXBGS)", "118.04", "Strong USD = EM headwind", "Elevated", "FRED"),
    ("Vietnam Interbank ON", "7.8%", "ELEVATED (+3.56bp WoW)", "Tight liquidity", "SBV"),
    ("Credit growth YoY", "1.0%", "VERY WEAK (-11pp WoW)", "Collapsing transmission", "SBV"),
    ("OMO net", "+19,160bn VND", "SBV injecting — easing", "Easing push", "SBV OMO"),
    ("SBV ref rate USD/VND", "25,137", "Stable", "No devaluation signal", "SBV"),
    ("US CPI YoY", "3.81%", "Persistent", "No Fed cut imminent", "BLS"),
    ("NFP MoM", "+115k", "Cooling", "Slower than expected", "BLS"),
    ("P(Fed cut 3m)", "35%", "Low", "FOMC on hold", "Model estimate"),
    ("P(VN tightening 1m)", "25%", "Moderate risk", "SBV liquidity squeeze possible", "Model estimate"),
]
for row in macro_data:
    ws7.append(list(row))
    r = ws7.max_row
    sig = str(row[2]).upper()
    if any(x in sig for x in ["ELEVATED", "TIGHT", "WEAK", "NEGATIVE", "STRONG USD"]):
        ws7.cell(r, 3).fill = REJECT_FILL
    elif any(x in sig for x in ["POSITIVE", "EASING", "STABLE", "NEUTRAL"]):
        ws7.cell(r, 3).fill = CONFIRM_FILL
    else:
        ws7.cell(r, 3).fill = REVISED_FILL

ws7.append([])
ws7.append(["REGIME CONCLUSION", "", "", "", ""])
ws7.cell(ws7.max_row, 1).font = BOLD

conclusions = [
    ("Classification", "FRAGILE UPTREND / Distribution Risk (State B)"),
    ("Primary evidence", "5 distribution days; VN30 below MA20; interbank ON 7.8%; credit growth 1%"),
    ("Bull case", "FTSE upgrade Sep 2026; OMO easing; earnings growth 15-18%"),
    ("Bear case", "UST 10Y > 5%; VND weakens; distribution day cluster continues"),
    ("1-3 month regime shift probability", "30% risk-off / 40% sideways / 30% moderate bounce"),
    ("Recommended portfolio stance", "SECTOR ROTATION — banks + quality growth; reduce high-beta RE/aviation"),
]
for row in conclusions:
    ws7.append(list(row))
    ws7.cell(ws7.max_row, 1).font = BOLD
    ws7.cell(ws7.max_row, 2).fill = REVISED_FILL

auto_width(ws7)

# -----------------------------------------------------------------------
# TAB 8: SOURCE LOG
# -----------------------------------------------------------------------
ws8 = wb.create_sheet("08_Source_Log")
ws8.append(["Source", "Type", "Date/Period", "Key Data Extracted", "Verified?"])
ws8.cell(1,1).font = HDR_FONT
ws8.cell(1,1).fill = HEADER_FILL
for c in range(1, 6):
    style_header(ws8.cell(1, c))

sources = [
    ("FireAnt SSOT fa_quarterly.parquet", "Primary FA", "2016Q1-2026Q2", "LTM profits, book equity, BVPS, shares", "YES — local SSOT"),
    ("FireAnt SSOT ta_ohlcv_panel.parquet", "Primary TA", "2017-05-18 to 2026-05-15", "Daily OHLCV prices", "YES — local SSOT"),
    ("FireAnt SSOT ta_vnindex.parquet", "VN-Index TA", "2012-01-03 to 2026-05-15", "VN-Index daily close", "YES — local SSOT"),
    ("TradingView screener (Vietnam)", "Market data", "2026-05-29", "P/E, P/B, market cap, sector (top 200)", "YES — real-time"),
    ("Weekly report (local)", "Macro/regime", "2026-05-29", "VN-Index, interbank, OMO, UST, DXY, credit", "YES — automated"),
    ("FRED (DGS2, DGS10, DTWEXBGS)", "Macro", "2026-05-14", "UST 2Y/10Y, USD broad", "YES — FRED API"),
    ("SSI 2026 Market Outlook", "Broker research", "Oct 2025", "Forward P/E target 12x at 1,920", "FACT from report"),
    ("Edison Group VNH report", "External research", "2026", "Forward P/E 12.7x; EPS growth 18%", "CITED — not primary"),
    ("SimplyWallSt HOSE analysis", "Aggregator", "2026-05-29", "HOSE P/E 11.2x (relative), 14.3x absolute", "REFERENCE — methodology unclear"),
    ("VCB MBS update report", "Broker research", "Jun 2025", "VCB P/B target 2.3x", "CITED — not primary"),
    ("MBS Banking 2026 Outlook", "Broker research", "Dec 2025", "Banking sector EPS growth 17%+", "CITED — not primary"),
    ("VinaCapital 2026 Outlook", "Fund manager", "Jan 2026", "Forward P/E 12x; earnings growth 14.5%", "CITED — not primary"),
]
for row in sources:
    ws8.append(list(row))

auto_width(ws8)

# -----------------------------------------------------------------------
# TAB 9: OPEN ISSUES
# -----------------------------------------------------------------------
ws9 = wb.create_sheet("09_Open_Issues")
ws9.append(["Open Issues & Data Gaps", "", "", ""])
ws9.cell(1,1).font = HDR_FONT
ws9.cell(1,1).fill = HEADER_FILL
ws9.merge_cells('A1:D1')

ws9.append(["Issue", "Impact", "What is Needed", "Priority"])
for c in range(1, 5):
    style_sub(ws9.cell(2, c))

issues = [
    ("2009, 2012, 2016 historical P/E unverifiable", "HIGH — cannot verify early cycle bottom claims", "FiinPro historical P/E series or HOSE official index P/E data going back to 2009", "HIGH"),
    ("Forward 2026F EPS consensus unavailable", "HIGH — cannot directly verify forward P/E claims", "Bloomberg or FiinPro consensus EPS for VN-Index constituents", "HIGH"),
    ("ICB codes null in FA quarterly parquet", "MEDIUM — cannot break financials sub-sector by ICB precisely", "Re-fetch FA data with ICB field populated; or maintain separate ICB-to-ticker mapping", "MEDIUM"),
    ("VN-Index official constituent list missing", "MEDIUM — using all HOSE stocks as proxy, not exact VN-Index weights", "HOSE official VN-Index constituent list and weights", "MEDIUM"),
    ("ACV P/E null in TradingView", "LOW — ACV is UPCOM listed; TV coverage incomplete", "Manual fetch from ACV quarterly reports", "LOW"),
    ("VPL P/E null — new listing", "LOW — only 2 quarters of data", "Wait 2-3 more quarters for TTM history", "LOW"),
    ("STB earnings normalization timeline", "MEDIUM — STB P/E 39x is distorted", "VAMC resolution schedule; STB 2025 annual report clean earnings", "MEDIUM"),
    ("BVPS unit consistency check at cycle bottoms", "MEDIUM — P/B historical may be slightly off if BVPS units changed", "Cross-check BVPS for 2019Q4 vs known VCB book equity at that date", "LOW"),
    ("Small-cap distortion in full-market P/E (14.4x vs 13.25x top-200)", "MEDIUM — which is more representative of VN-Index?", "Use official VN-Index P/E from HOSE vs our calculation", "MEDIUM"),
    ("Forward P/E coverage ratio by market cap", "MEDIUM — broker consensus may not cover all VN-Index stocks", "FiinPro consensus coverage disclosure", "LOW"),
]
for row in issues:
    ws9.append(list(row))
    r = ws9.max_row
    priority = str(row[3]).upper()
    if "HIGH" in priority:
        ws9.cell(r, 4).fill = REJECT_FILL
    elif "MEDIUM" in priority:
        ws9.cell(r, 4).fill = REVISED_FILL
    else:
        ws9.cell(r, 4).fill = CONFIRM_FILL

auto_width(ws9)

# -----------------------------------------------------------------------
# SAVE
# -----------------------------------------------------------------------
wb.save(OUTPUT_PATH)
print(f"Excel saved: {OUTPUT_PATH}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
